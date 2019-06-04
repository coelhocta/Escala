import openpyxl
from datetime import date


def data_num(d):
    num = date.toordinal(d)
    return num


def num_data(d):
    num = date.fromordinal(d)
    return num


def gera_nomes():
    tmp = {}
    indisp = []
    lin = aba_inicio.max_row
    col = aba_inicio.max_column
    for i in range(8, lin - 8):
        tmp['Antig'] = i -8
        tmp['Nome'] = aba_inicio.cell(i, 1).value
        for c in range(2, aba_inicio.max_column):
            d = aba_inicio.cell(i, c).value
            if d != None:
                e = data_num(d)
                indisp.append(e)
        tmp['Indisp'] = indisp.copy()
        indisp.clear()
        nomes.append(tmp.copy())
        tmp.clear()
    return nomes


def gera_periodo():
    i = data_num(aba_inicio['B1'].value)
    f = data_num(aba_inicio['C1'].value)
    while i <= f:
        periodo.append(i)
        i += 1


def gera_quadrinho():
    # Busca Roxa da Planilha
    for r in aba_inicio['B3':'AZ3']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                roxa.append(data)

    # Busca Vermelha da Planilha
    for r in aba_inicio['B4':'AZ4']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                vermelha.append(data)

    # Busca Marrom da Planilha
    for r in aba_inicio['C5':'AZ5']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                marrom.append(data)

    # Gera vermelha e Preta Automática
    for d in periodo:
        if date.weekday(num_data(d)) in (5, 6) and d not in vermelha and d not in roxa:
            vermelha.append(d)
        vermelha.sort()

    # Gera Marrom Automática
    for d in vermelha:
        if (d -1) not in vermelha and (d -1) not in roxa:
            marrom.append(d - 1)
        for d in roxa:
            if (d - 1) not in roxa:
                marrom.append(d - 1)
        marrom.sort()

    # Gera Preta Automática
    for d in periodo:
        if d not in vermelha and d not in roxa and d not in marrom:
            preta.append(d)


def fila(cor):
    fila = []
    z = {'roxa':lastro_roxa, 'vermelha':lastro_vermelha, 'marrom':lastro_marrom, 'preta':lastro_preta}
    for a in z[cor]:
        b = [a[0], a[1]]
        fila.append(b)
    fila.reverse()
    fila = sorted(fila, key=lambda x: x[1])
    return fila



def busca_lastro_planilha():
    tmp = []
    tmp1 = []
    # Lastro Roxa
    for i in range(3, aba_rox.max_row + 1):
        for j in range(1,(aba_rox.max_column)+1):
            conteudo = aba_rox.cell(row=i, column=j).value
            if conteudo != None:
                if type(conteudo) is not str:
                    conteudo = data_num(conteudo)
                tmp.append(conteudo)
        tmp1.append(tmp.copy())
        tmp.clear()
    for a in tmp1:
        b = [a, len(a)]
        lastro_roxa.append(b)
    # Lastro Vermelha
    for i in range(3, aba_ver.max_row + 1):
        for j in range(1,(aba_ver.max_column)+1):
            conteudo = aba_ver.cell(row=i, column=j).value
            if conteudo != None:
                if type(conteudo) is not str:
                    conteudo = data_num(conteudo)
                tmp.append(conteudo)
        tmp1.append(tmp.copy())
        tmp.clear()
    for a in tmp1:
        b = [a, len(a)]
        lastro_vermelha.append(b)
    # Lastro Marrom
    for i in range(3, aba_mar.max_row + 1):
        for j in range(1, (aba_mar.max_column) + 1):
            conteudo = aba_mar.cell(row=i, column=j).value
            if conteudo != None:
                if type(conteudo) is not str:
                    conteudo = data_num(conteudo)
                tmp.append(conteudo)
        tmp1.append(tmp.copy())
        tmp.clear()
    for a in tmp1:
        b = [a, len(a)]
        lastro_marrom.append(b)
    # Lastro Preta
    for i in range(3, aba_pre.max_row + 1):
        for j in range(1,(aba_pre.max_column)+1):
            conteudo = aba_pre.cell(row=i, column=j).value
            if conteudo != None:
                if type(conteudo) is not str:
                    conteudo = data_num(conteudo)
                tmp.append(conteudo)
        tmp1.append(tmp.copy())
        tmp.clear()
    for a in tmp1:
        b = [a, len(a)]
        lastro_preta.append(b)





##########################################
# Lê o nome e as abas da planilha


wb = openpyxl.load_workbook('Escala.xlsx')
aba_inicio = wb['Inicio']
aba_ver = wb['Vermelha']
aba_pre = wb['Preta']
aba_mar = wb['Marrom']
aba_rox = wb['Roxa']

##########################################
# Gera os dias da Semana
diaSemana = ['SEGUNDA-FEIRA', 'TERÇA-FEIRA', 'QUARTA-FEIRA', 'QUINTA-FEIRA', 'SEXTA-FEIRA', 'SÁBADO', 'DOMINGO']

##########################################
# Listas
nomes = []
periodo = []
vermelha = []
marrom = []
roxa = []
preta = []
lastro_roxa = []
lastro_vermelha = []
lastro_marrom = []
lastro_preta = []

##########################################
# Chamadas Funções
gera_nomes()
gera_periodo()
gera_quadrinho()
busca_lastro_planilha()

##########################################
# Gerar lista sequencia roxa

print(fila('roxa'))
lastro_roxa[25][1] += 1
lastro_roxa[31][1] += 1
print(fila('vermelha'))
print(fila('marrom'))
print(fila('preta'))

'''
print(f'Nomes: {nomes}')
print(f'Período: {periodo}')
print(f'Roxa: {roxa}')
print(f'Vermelha: {vermelha}')
print(f'Marrom: {marrom}')
print(f'Preta: {preta}')

for a in periodo:
    if a in roxa:
        cor = 'Roxa'
    if a in vermelha:
        cor = 'Vermelha'
    if a in marrom:
        cor = 'Marrom'
    if a in preta:
        cor = 'Preta'
    print(f'{cor:>8} - {diaSemana[date.weekday(num_data(a))]:^13} - {num_data(a)}')


a1 = aba_mil_Ind['A1']
a2 = aba_mil_Ind['A2']
a3 = aba_mil_Ind.cell(3, 1)

print(a1.value)
print(a2.value)
print(a3.value)

# mostra o número de linhas da planilha aba_mil_ind
print(aba_mil_ind.max_row)

# mostra o número de Colunas da planilha aba_mil_ind
print(aba_mil_ind.max_column)

# mostra o conteúdo da célula A1 até B10
for r in aba_mil_ind['A1':'B10']:
    for c in r:
        print(c.value)

# Converte data da planilha
import datetime
ws['A2'] = datetime.datetime.now()

# Mostra as células A3 até C10
for c1, c2, c3 in aba_rox['A3': 'C10']:
    print("{} {} {}".format(c1.value, c2.value, c3.value))


#####################################################

# Edita a nova planilha
planilha = openpyxl.Workbook()
aba =  wb['Militares.Indisponibilidades'] # Seleciona a aba a ser modificada
aba.title = 'Aba1' # Altera o nome da aba selecionada
aba['A3'].value = 'Teste de escrita' # Escreve na célula escolhida
aba.create_sheet(title='NovaAba' # Cria uma nova aba
wb.save('Escala.final.xlsx') # Salva uma cópia com o novo nome ad planilha
'''