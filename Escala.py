import openpyxl
from datetime import date
from openpyxl.styles import colors, Font, Alignment
from openpyxl.styles.borders import Border, Side


def data_num(d):
    num = date.toordinal(d)
    return num


def num_data(d):
    num = date.fromordinal(d)
    return num


def gera_nomes():
    tmp1 = {}
    indisp = []
    lin = aba_inicio.max_row
    col = aba_inicio.max_column
    for i in range(8, lin - 8):
        tmp1['antig'] = i - 8
        tmp1['nome'] = aba_inicio.cell(i, 1).value
        for c1 in range(2, col):
            d = aba_inicio.cell(i, c1).value
            if d is not None:
                e = data_num(d)
                indisp.append(e)
        tmp1['indisp'] = indisp.copy()
        indisp.clear()
        nomes.append(tmp1.copy())
        tmp1.clear()
    return nomes


def gera_periodo():
    i = data_num(aba_inicio['B1'].value)
    f = data_num(aba_inicio['C1'].value)
    while i <= f:
        periodo.append(i)
        i += 1


def gera_quadrinho():
    # Busca da Planilha
    for t in cores:
        for r in t['dias']:
            for c2 in r:
                if c2.value is not None:
                    data = data_num(c2.value)
                    if data in periodo:
                        t['cor'].append(data)

    # Gera Vermelha e Marrom Automática
    for d in periodo:
        if date.weekday(num_data(d)) in (5, 6) and d not in vermelha and d not in roxa:
            vermelha.append(d)
        if date.weekday(num_data(d)) is 4 and d not in vermelha and d not in roxa:
            marrom.append(d)
        if date.weekday(num_data(d)) is 4 and d not in marrom:
            a1 = d
            while a1 not in marrom:
                if a1 not in vermelha and a1 not in roxa:
                    marrom.append(a1)
                    break
                a1 -= 1

    # # Gera Marrom Automática caso seja antes de qualquer vermelha.
    # for d in vermelha:
    #     dia = d - 1
    #     if dia not in vermelha and dia not in roxa and dia in periodo:
    #         marrom.append(dia)
    # for d in roxa:
    #     dia = d - 1
    #     if dia not in roxa and dia in periodo:
    #         marrom.append(dia)
    # Gera Preta Automática
    for d in periodo:
        if d not in vermelha and d not in roxa and d not in marrom:
            preta.append(d)


def fila_ver():
    fila = []
    for a1 in lastro_vermelha:
        c3 = [len(a1['lastros']), a1['antig'], a1['nome']]
        fila.append(c3)
    fila.reverse()
    fila = sorted(fila, key=lambda x: x[0])
    return fila


def fila_mar():
    fila = []
    for a2 in lastro_marrom:
        if "*" not in a2['nome']:
            c4 = [len(a2['lastros']), a2['antig'], a2['nome']]
            fila.append(c4)
    fila.reverse()
    fila = sorted(fila, key=lambda x: x[0])
    return fila


def fila_pre():
    fila = []
    for a4 in lastro_preta:
        if "*" not in a4['nome']:
            c6 = [len(a4['lastros']), a4['antig'], a4['nome']]
            fila.append(c6)
    fila.reverse()
    fila = sorted(fila, key=lambda x: x[0])
    return fila


def busca_lastro_planilha():
    tmp2 = {}
    tmp1 = []
    for a5 in cores:
        for i in range(2, a5['linhas'] + 1):
            tmp2['cor'] = a5['cor_texto']
            tmp2['antig'] = i - 2
            tmp2['nome'] = a5['conteudo'](row=i, column=1).value
            for j in range(1, (a5['colunas'])+1):
                conteudo = a5['conteudo'](row=i, column=j+1).value
                if conteudo is not None:
                    if type(conteudo) is not str:
                        conteudo = data_num(conteudo)
                    tmp1.append(conteudo)
            tmp2['lastros'] = tmp1.copy()
            a5['lastro'].append(tmp2.copy())
            tmp1.clear()
        tmp2.clear()


def preenche_from_planilha():
    # Busca escala forçada da planilha
    tmp5 = {}
    for f in cores:
        for a7 in f['lastro']:
            for b4 in a7['lastros']:
                if b4 in f['cor'] and b4 in periodo:
                    tmp5['cor'] = f['cor_texto']
                    tmp5['diaSemana'] = diaSemana[date.weekday(num_data(b4))]
                    tmp5['dia'] = b4
                    tmp5['nome'] = a7['nome']
                    tmp5['antig'] = a7['antig']
                    escala_final.append(tmp5.copy())
                    tmp5.clear()


def fila_ver_reserva():
    fila = []
    for a9 in lastro_vermelha_reserva:
        c9 = [a9['lastro'], a9['antig'], a9['nome']]
        fila.append(c9)
    fila.reverse()
    fila = sorted(fila, key=lambda x: x[0])
    return fila


wb = openpyxl.load_workbook('Escala.xlsx')
wb.remove(wb['Escala'])
wb.create_sheet('Escala')
aba_inicio = wb['Inicio']
aba_ver = wb['Vermelha']
aba_pre = wb['Preta']
aba_mar = wb['Marrom']
aba_rox = wb['Roxa']
aba_escala = wb['Escala']

# Listas
nomes = []
periodo = []
vermelha = []
marrom = []
roxa = []
preta = []
lastro_roxa = []
lastro_vermelha = []
lastro_marrom = []
lastro_preta = []
escala_final = []
diaSemana = ['SEGUNDA-FEIRA', 'TERÇA-FEIRA', 'QUARTA-FEIRA', 'QUINTA-FEIRA', 'SEXTA-FEIRA', 'SÁBADO', 'DOMINGO']
cores = [{'cor_texto': 'ROXA', 'dias': aba_inicio['B3':'AZ3'], 'cor':roxa, 'linhas': aba_rox.max_row,
          'colunas':aba_rox.max_column, 'conteudo': aba_rox.cell, 'lastro':lastro_roxa},
         {'cor_texto': 'VERMELHA', 'dias': aba_inicio['B4':'AZ4'], 'cor':vermelha, 'linhas': aba_ver.max_row,
          'colunas':aba_ver.max_column, 'conteudo': aba_ver.cell, 'lastro':lastro_vermelha},
         {'cor_texto': 'MARROM', 'dias': aba_inicio['B5':'AZ5'], 'cor':marrom, 'linhas': aba_mar.max_row,
          'colunas':aba_mar.max_column, 'conteudo': aba_mar.cell, 'lastro':lastro_marrom},
         {'cor_texto': 'PRETA', 'dias': aba_inicio['B6':'AZ6'], 'cor':preta, 'linhas': aba_pre.max_row,
          'colunas':aba_pre.max_column, 'conteudo': aba_pre.cell, 'lastro':lastro_preta}]

##########################################
# Chamadas Funções
gera_nomes()
gera_periodo()
gera_quadrinho()
busca_lastro_planilha()
preenche_from_planilha()

##########################################
for a in nomes:
    a['lastro_total'] = []
    for b in lastro_roxa:
        if a['antig'] == (b['antig']):
            for c in b['lastros']:
                a['lastro_total'].append(c)
    for b in lastro_vermelha:
        if a['antig'] == (b['antig']):
            for c in b['lastros']:
                a['lastro_total'].append(c)
    for b in lastro_marrom:
        if a['antig'] == (b['antig']):
            for c in b['lastros']:
                a['lastro_total'].append(c)
    for b in lastro_preta:
        if a['antig'] == (b['antig']):
            for c in b['lastros']:
                a['lastro_total'].append(c)

vermelha_copy = sorted(vermelha.copy())
marrom_copy = sorted(marrom.copy())
preta_copy = sorted(preta.copy())

for a in escala_final:
    if a['dia'] in preta_copy:
        preta_copy.remove(a['dia'])
    if a['dia'] in vermelha_copy:
        vermelha_copy.remove(a['dia'])
    if a['dia'] in marrom_copy:
        marrom_copy.remove(a['dia'])

cont = 0
for a in marrom_copy:
    fila_marrom = fila_mar()
    tmp = {'cor': 'MARROM', 'diaSemana': diaSemana[date.weekday(num_data(a))], 'dia': a, 'nome': ''}
    while True:
        for b in nomes:
            if b['antig'] == fila_marrom[cont][1]:
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' not in b['nome']:
                    tmp['nome'] = fila_marrom[cont][2]
                    tmp['antig'] = fila_marrom[cont][1]
                    lastro_marrom[b['antig']]['lastros'].append(a)
                    b['lastro_total'].append(a)
                    escala_final.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                else:
                    cont += 1
        if not tmp:
            break

cont = 0
for a in vermelha_copy:
    fila_vermelha = fila_ver()
    tmp = {'cor': 'VERMELHA', 'diaSemana': diaSemana[date.weekday(num_data(a))], 'dia': a, 'nome': ''}
    while True:
        for b in nomes:
            if b['antig'] == fila_vermelha[cont][1]:
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' in b['nome']\
                        and a + 1 not in preta\
                        and a + 1 not in marrom:
                    tmp['nome'] = fila_vermelha[cont][2]
                    tmp['antig'] = fila_vermelha[cont][1]
                    lastro_vermelha[b['antig']]['lastros'].append(a)
                    b['lastro_total'].append(a)
                    escala_final.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' not in b['nome']:
                    tmp['nome'] = fila_vermelha[cont][2]
                    tmp['antig'] = fila_vermelha[cont][1]
                    lastro_vermelha[b['antig']]['lastros'].append(a)
                    b['lastro_total'].append(a)
                    escala_final.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                else:
                    cont += 1
        if not tmp:
            break

cont = 0
for a in preta_copy:
    fila_preta = fila_pre()
    tmp = {'cor': 'PRETA', 'diaSemana': diaSemana[date.weekday(num_data(a))], 'dia': a, 'nome': ''}
    while True:
        for b in nomes:
            if b['antig'] == fila_preta[cont][1]:
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' not in b['nome']:
                    tmp['nome'] = fila_preta[cont][2]
                    tmp['antig'] = fila_preta[cont][1]
                    lastro_preta[b['antig']]['lastros'].append(a)
                    b['lastro_total'].append(a)
                    escala_final.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                else:
                    cont += 1
        if not tmp:
            break

meses = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho', 'julho',
         'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
nomePlanilha = (meses[num_data(periodo[0]-1).month])

titulo = f'Escala de serviço para o mês de {nomePlanilha}.'

escala_planilha = [(titulo, ''), ('Data', 'Dia da Semana', 'Militar', 'Cor', 'OBS:')]
for a in periodo:
    for b in escala_final:
        if a == b['dia']:
            dia = date.strftime(num_data(b["dia"]), "%d/%m/%Y")
            tmp = (dia, str(b["diaSemana"]), str(b["nome"]), str(b['cor']))
            escala_planilha.append(tmp)

for a in escala_planilha:
    aba_escala.append(a)

# Coloca cor e borda nas células
for l1, a in enumerate(aba_escala):
    for b in range(len(a)):
        if a[b].value == 'VERMELHA':
            a[b].font = Font(color=colors.RED, bold=True)
            a[b-1].font = Font(color=colors.RED, bold=True)
            a[b - 2].font = Font(color=colors.RED, bold=True)
            a[b - 3].font = Font(color=colors.RED, bold=True)
        if a[b].value == 'ROXA':
            a[b].font = Font(color='800080', bold=True)
            a[b - 1].font = Font(color='800080', bold=True)
            a[b - 2].font = Font(color='800080', bold=True)
            a[b - 3].font = Font(color='800080', bold=True)
        if a[b].value == 'MARROM':
            a[b].font = Font(color='8b4513', bold=True)
            a[b - 1].font = Font(color='8b4513', bold=True)
            a[b - 2].font = Font(color='8b4513', bold=True)
            a[b - 3].font = Font(color='8b4513', bold=True)
        if a[b].value == 'PRETA':
            a[b].font = Font(bold=True)
            a[b - 1].font = Font(bold=True)
            a[b - 2].font = Font(bold=True)
            a[b - 3].font = Font(bold=True)
        else:
            a[b].font = Font(bold=True)
        a[b].alignment = Alignment(horizontal='center')
        if l1 > 0 and b < 5:
            a[b].border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                 bottom=Side(style='medium'))
        if l1 == 0:
            a[b].font = Font(bold=True, size=15, name='Arial')

# Apaga a coluna com o texto cores
aba_escala.delete_cols(4)

# Redimensiona o tamanho das colunas
aba_escala.column_dimensions['A'].width = 14
aba_escala.column_dimensions['B'].width = 26
aba_escala.column_dimensions['C'].width = 26
aba_escala.column_dimensions['D'].width = 26


#############################
'''
'# Inserir imagem
from openpyxl.drawing.image import Image
img = Image('logo.png')
aba_escala.add_image(img, 'A1')
'''
##########################

aba_ver.delete_rows(2, aba_ver.max_row)
for a in lastro_vermelha:
    temp = list()
    temp.append(a['nome'])
    for b in a['lastros']:
        if type(b) is int:
            temp.append(num_data(b))
        else:
            temp.append(b)
    aba_ver.append(temp)
    temp.clear()
# Coloca cor e borda nas aba Vermelha
for l, a in enumerate(aba_ver):
    for b in range(len(a)):
        if a[b].value is not None and l > 0:
            a[b].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            a[b].font = Font(bold=True)
            a[b].number_format = 'dd/mm/yyyy'

aba_mar.delete_rows(2, aba_mar.max_row)
for a in lastro_marrom:
    temp = list()
    temp.append(a['nome'])
    for b in a['lastros']:
        if type(b) is int:
            temp.append(num_data(b))
        else:
            temp.append(b)
    aba_mar.append(temp)
    temp.clear()
# Coloca cor e borda nas aba Marrom
for l, a in enumerate(aba_mar):
    for b in range(len(a)):
        if a[b].value is not None and l > 0:
            a[b].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            a[b].font = Font(bold=True)
            a[b].number_format = 'dd/mm/yyyy'

temp = []
aba_pre.delete_rows(2, aba_pre.max_row)
for a in lastro_preta:
    temp.append(a['nome'])
    for b in a['lastros']:
        if type(b) is int:
            temp.append(num_data(b))
        else:
            temp.append(b)
    aba_pre.append(temp)
    temp.clear()
# Coloca cor e borda nas aba Preta
for l, a in enumerate(aba_pre):
    for b in range(len(a)):
        if a[b].value is not None and l > 0:
            a[b].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            a[b].font = Font(bold=True)
            a[b].number_format = 'dd/mm/yyyy'

aba_rox.delete_rows(2, aba_rox.max_row)
for a in lastro_roxa:
    temp.append(a['nome'])
    for b in a['lastros']:
        if type(b) is int:
            temp.append(num_data(b))
        else:
            temp.append(b)
    aba_rox.append(temp)
    temp.clear()
# Coloca cor e borda nas aba Roxa
for l, a in enumerate(aba_rox):
    for b in range(len(a)):
        if a[b].value is not None and l > 0:
            a[b].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            a[b].font = Font(bold=True)
            a[b].number_format = 'dd/mm/yyyy'

##########################
# Reservas

lastro_vermelha_reserva = []
escala_reserva_vermelha = []
for a in lastro_vermelha:
    b = {'antig': a['antig'], 'nome': a['nome'], 'lastro': len(a['lastros'])}
    lastro_vermelha_reserva.append(b)

cont = 0
for a in sorted(vermelha):
    fila_vermelha_reserva = fila_ver_reserva()
    tmp = {'cor': 'VERMELHA', 'dia': a, 'nome': ''}
    while True:
        for b in nomes:
            if b['antig'] == fila_vermelha_reserva[cont][1]:
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' in b['nome']\
                        and a + 1 not in preta\
                        and a + 1 not in marrom:
                    tmp['nome'] = fila_vermelha_reserva[cont][2]
                    tmp['antig'] = fila_vermelha_reserva[cont][1]
                    lastro_vermelha_reserva[b['antig']]['lastro'] += 1
                    b['lastro_total'].append(a)
                    escala_reserva_vermelha.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                if a not in b['lastro_total'] \
                        and a - 1 not in (b['lastro_total']) \
                        and a + 1 not in (b['lastro_total']) \
                        and a + 2 not in (b['lastro_total']) \
                        and a - 2 not in (b['lastro_total'])\
                        and a not in b['indisp']\
                        and '*' not in b['nome']:
                    tmp['nome'] = fila_vermelha_reserva[cont][2]
                    tmp['antig'] = fila_vermelha_reserva[cont][1]
                    lastro_vermelha_reserva[b['antig']]['lastro'] += 1
                    b['lastro_total'].append(a)
                    escala_reserva_vermelha.append(tmp.copy())
                    tmp.clear()
                    cont = 0
                    break
                else:
                    cont += 1
        if not tmp:
            break

fila_marrom = fila_mar()
fila_preta = fila_pre()
escala_planilha_reserva_vermelha = [(), ('RESERVAS:', ''), ('VERMELHA', '', 'MARROM', 'PRETA')]
for a in periodo:
    for l, b in enumerate(escala_reserva_vermelha):
        if a == b['dia']:
            if l < 3:
                mar = fila_marrom[l][2]
                pre = fila_preta[l][2]
            else:
                mar = ''
                pre = ''
            dia = date.strftime(num_data(b["dia"]), "%d/%m/%Y")
            tmp = (dia, str(b["nome"]), mar, pre)
            escala_planilha_reserva_vermelha.append(tmp)

for a in escala_planilha_reserva_vermelha:
    aba_escala.append(a)

# Coloca cor e borda
for l, a in enumerate(aba_escala):
    if l > len(escala_planilha):
        for b in range(len(a)):
            if a[b].value is not None and l > 0 and not a[b].value == '':
                a[b].border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                     bottom=Side(style='medium'))
                a[b].font = Font(bold=True)
                a[b].alignment = Alignment(horizontal='center')
            if b < 2 and l > len(escala_planilha)+1:
                a[b].font = Font(color=colors.RED, bold=True)
            if b == 2:
                a[b].font = Font(color='8b4513', bold=True)


aba_escala.merge_cells(start_row=len(escala_planilha)+2, start_column=1, end_row=len(escala_planilha)+2, end_column=4)
aba_escala.merge_cells(start_row=len(escala_planilha)+3, start_column=1, end_row=len(escala_planilha)+3, end_column=2)
aba_escala.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

aba_escala.page_margins.top = 0.5
aba_escala.page_margins.bottom = 0.5
aba_escala.page_margins.left = 0.5
aba_escala.page_margins.right = 0.5
aba_escala.page_horizontalCentered = True
aba_escala.page_verticalCentered = True

assinaura1 = aba_inicio['F1'].value
assinaura2 = aba_inicio['F2'].value
assinaura3 = aba_inicio['J1'].value
assinaura4 = aba_inicio['J2'].value


dataAssinatura = 'São José dos Campos, ' + str(date.today().day) + ' de ' + \
                 str(meses[date.today().month-1]) + ' de ' + str(date.today().year)
dataRodape = [(dataAssinatura, ''),(), ('_____________________________________','','_____________________________________'),
              (assinaura1, '', assinaura3), (assinaura2, '', assinaura4)]
for a in dataRodape:
    aba_escala.append(a)

ultimaLinha = len(escala_planilha)+len(escala_reserva_vermelha)


aba_escala.merge_cells(start_row=ultimaLinha+4, start_column=1, end_row=ultimaLinha+4, end_column=2)
aba_escala.merge_cells(start_row=ultimaLinha+6, start_column=1, end_row=ultimaLinha+6, end_column=2)
aba_escala.merge_cells(start_row=ultimaLinha+6, start_column=3, end_row=ultimaLinha+6, end_column=4)
aba_escala.merge_cells(start_row=ultimaLinha+7, start_column=1, end_row=ultimaLinha+7, end_column=2)
aba_escala.merge_cells(start_row=ultimaLinha+7, start_column=3, end_row=ultimaLinha+7, end_column=4)
aba_escala.merge_cells(start_row=ultimaLinha+8, start_column=1, end_row=ultimaLinha+8, end_column=2)
aba_escala.merge_cells(start_row=ultimaLinha+8, start_column=3, end_row=ultimaLinha+8, end_column=4)

aba_escala.cell(row=ultimaLinha+8, column=1).font = Font(bold=True)
for linha in range(ultimaLinha+4,ultimaLinha+9):
    for coluna in range(1,5):
        aba_escala.cell(row=linha, column=coluna).font = Font(bold=True)
        aba_escala.cell(row=linha, column=coluna).alignment = Alignment(horizontal='center')

##########################

wb.save(f'Escala.{nomePlanilha}.xlsx')
