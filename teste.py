import openpyxl
from datetime import date


def data_num(d):
    num = date.toordinal(d)
    return num


def num_data(d):
    num = date.fromordinal(d)
    return num


def gera_nomes():
    tmp = {}
    indisp = []
    lin = aba_inicio.max_row
    col = aba_inicio.max_column
    for i in range(8, lin - 8):
        tmp['Antig'] = i -8
        tmp['Nome'] = aba_inicio.cell(i, 1).value
        for c in range(2, aba_inicio.max_column):
            d = aba_inicio.cell(i, c).value
            if d != None:
                e = data_num(d)
                indisp.append(e)
        tmp['Indisp'] = indisp.copy()
        indisp.clear()
        nomes.append(tmp.copy())
        tmp.clear()
    return nomes

nomes = []
wb = openpyxl.load_workbook('Escala.xlsx')
aba_inicio = wb['Inicio']
print(nomes)

def gera_periodo():
    i = data_num(aba_inicio['B1'].value)
    f = data_num(aba_inicio['C1'].value)
    while i <= f:
        periodo.append(i)
        i += 1


def gera_quadrinho():
    # Busca Roxa da Planilha
    for r in aba_inicio['B3':'AZ3']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                roxa.append(data)

    # Busca Vermelha da Planilha
    for r in aba_inicio['B4':'AZ4']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                vermelha.append(data)

    # Busca Marrom da Planilha
    for r in aba_inicio['C5':'AZ5']:
        for c in r:
            if c.value != None:
                data = data_num(c.value)
                marrom.append(data)

    # Gera vermelha e Preta Automática
    for d in periodo:
        if date.weekday(num_data(d)) in (5, 6) and d not in vermelha and d not in roxa:
            vermelha.append(d)
        vermelha.sort()

    # Gera Marrom Automática
    for d in vermelha:
        if (d -1) not in vermelha and (d -1) not in roxa:
            marrom.append(d - 1)
        for d in roxa:
            if (d - 1) not in roxa:
                marrom.append(d - 1)
        marrom.sort()

    # Gera Preta Automática
    for d in periodo:
        if d not in vermelha and d not in roxa and d not in marrom:
            preta.append(d)


##########################################
# Lê o nome e as abas da planilha



aba_ver = wb['Vermelha']
aba_pre = wb['Preta']
aba_mar = wb['Marrom']
aba_rox = wb['Roxa']

##########################################
# Gera os dias da Semana
diaSemana = ['SEGUNDA-FEIRA', 'TERÇA-FEIRA', 'QUARTA-FEIRA', 'QUINTA-FEIRA', 'SEXTA-FEIRA', 'SÁBADO', 'DOMINGO']

##########################################
# Listas

periodo = []
vermelha = []
marrom = []
roxa = []
preta = []
lastro_roxa = []
lastro_vermelha = []
lastro_marrom = []
lastro_preta = []

##########################################
# Chamadas Funções
gera_nomes()
gera_periodo()
gera_quadrinho()

##########################################
# Gerar lista sequencia roxa

tmp = []
fila_roxa = []
fila_vermelha = []
fila_marrom = []
fila_preta = []
for i in range(3, aba_rox.max_row + 1):
    for j in range(1,(aba_rox.max_column)+1):
        conteudo = aba_rox.cell(row=i, column=j).value
        if conteudo != None:
            if type(conteudo) is not str:
                conteudo = data_num(conteudo)
            tmp.append(conteudo)
    lastro_roxa.append(tmp.copy())
    tmp.clear()
for a in lastro_roxa:
    b = a[0], len(a)
    fila_roxa.append(b)
fila_roxa.reverse()
fila_roxa = sorted(fila_roxa, key=lambda x: x[1])
for a in fila_roxa:
    print(a)

contador = 0
roxa_final = []
for dia in roxa:
    while True:
        pessoa = fila_roxa[contador][0]
        tmp = [dia, pessoa]
        roxa_final.append(tmp.copy())
        contador += 1
        break


print(nomes)

'''
print(f'Nomes: {nomes}')
print(f'Período: {periodo}')
print(f'Roxa: {roxa}')
print(f'Vermelha: {vermelha}')
print(f'Marrom: {marrom}')
print(f'Preta: {preta}')

for a in periodo:
    if a in roxa:
        cor = 'Roxa'
    if a in vermelha:
        cor = 'Vermelha'
    if a in marrom:
        cor = 'Marrom'
    if a in preta:
        cor = 'Preta'
    print(f'{cor:>8} - {diaSemana[date.weekday(num_data(a))]:^13} - {num_data(a)}')


a1 = aba_mil_Ind['A1']
a2 = aba_mil_Ind['A2']
a3 = aba_mil_Ind.cell(3, 1)

print(a1.value)
print(a2.value)
print(a3.value)

# mostra o número de linhas da planilha aba_mil_ind
print(aba_mil_ind.max_row)

# mostra o número de Colunas da planilha aba_mil_ind
print(aba_mil_ind.max_column)

# mostra o conteúdo da célula A1 até B10
for r in aba_mil_ind['A1':'B10']:
    for c in r:
        print(c.value)

# Converte data da planilha
import datetime
ws['A2'] = datetime.datetime.now()

# Mostra as células A3 até C10
for c1, c2, c3 in aba_rox['A3': 'C10']:
    print("{} {} {}".format(c1.value, c2.value, c3.value))


#####################################################

# Edita a nova planilha
planilha = openpyxl.Workbook()
aba =  wb['Militares.Indisponibilidades'] # Seleciona a aba a ser modificada
aba.title = 'Aba1' # Altera o nome da aba selecionada
aba['A3'].value = 'Teste de escrita' # Escreve na célula escolhida
aba.create_sheet(title='NovaAba' # Cria uma nova aba
wb.save('Escala.final.xlsx') # Salva uma cópia com o novo nome ad planilha
'''